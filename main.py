import PyPDF2
import os
import re
import shutil  # Para mover archivos
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Función para extraer la información de la factura
def extract_factura_info(pdf_file_path):
    invoice_info = {}

    # Abrir el archivo PDF
    with open(pdf_file_path, 'rb') as file:
        pdf_reader = PyPDF2.PdfReader(file)
        text = ''

        for page_num in range(len(pdf_reader.pages)):
            page = pdf_reader.pages[page_num]
            text += page.extract_text() or ''

        # Expresiones regulares para extraer la información relevante
        invoice_number_pattern = r"Factura de Venta N°\s*(\d+)"
        seller_pattern = r"Vendedor:\s*(.+)"
        nit_pattern = r"NIT:\s*([\d.]+-\d+)"
        address_pattern = r"Dirección:\s*(.+)"
        client_pattern = r"Cliente:\s*(.+)"
        client_nit_pattern = r"NIT/CC:\s*(\d+)"
        # Ajustar la expresión regular de fecha al formato "19 de septiembre de 2024"
        date_pattern = r"(\d{1,2}\s*de\s*[a-zA-Z]+\s*de\s*\d{4})"

        # Extraer datos sin incluir el teléfono del vendedor
        invoice_info['Número de factura'] = re.search(invoice_number_pattern, text).group(1) if re.search(invoice_number_pattern, text) else None
        invoice_info['Vendedor'] = re.search(seller_pattern, text).group(1).split(',')[0] if re.search(seller_pattern, text) else None  # Excluir teléfono
        invoice_info['NIT'] = re.search(nit_pattern, text).group(1) if re.search(nit_pattern, text) else None
        invoice_info['Dirección Vendedor'] = re.search(address_pattern, text).group(1) if re.search(address_pattern, text) else None
        invoice_info['Cliente'] = re.search(client_pattern, text).group(1) if re.search(client_pattern, text) else None
        invoice_info['NIT/CC Cliente'] = re.search(client_nit_pattern, text).group(1) if re.search(client_nit_pattern, text) else None
        # Verificar que la fecha se extraiga correctamente
        invoice_info['Fecha'] = re.search(date_pattern, text).group(1) if re.search(date_pattern, text) else 'Fecha no encontrada'

        # Extraer solo la descripción de los productos
        product_lines_pattern = r"([A-Za-z\s]+)\s+\d+\s+[\d,]+\s+\d+%\s+[\d,]+"
        products = re.findall(product_lines_pattern, text)
        # Convertir la lista de productos en una cadena separada por comas
        invoice_info['Productos'] = ', '.join([prod.strip() for prod in products])  # Convertir lista a cadena

        # Extraer subtotal, IVA y total
        subtotal_pattern = r"Subtotal:\s*([\d,]+)\s*COP"
        iva_pattern = r"IVA \(\d+%\):\s*([\d,]+)\s*COP"
        total_pattern = r"Total:\s*([\d,]+)\s*COP"

        invoice_info['Subtotal'] = int(re.search(subtotal_pattern, text).group(1).replace(',', '')) if re.search(subtotal_pattern, text) else None
        invoice_info['IVA'] = int(re.search(iva_pattern, text).group(1).replace(',', '')) if re.search(iva_pattern, text) else None
        invoice_info['Total'] = int(re.search(total_pattern, text).group(1).replace(',', '')) if re.search(total_pattern, text) else None

    return invoice_info

# Función para obtener los archivos PDF en una carpeta
def get_files_in_carpetas(folder_path):
    files = []
    for root, dirs, filenames in os.walk(folder_path):
        for filename in filenames:
            if filename.endswith('.pdf'):  # Asegúrate de procesar solo archivos PDF
                files.append(os.path.join(root, filename))
    return files

# Función para guardar los datos en Excel sin sobrescribir los anteriores
def save_to_excel(df, excel_file):
    if os.path.exists(excel_file):
        try:
            # Cargar el archivo existente
            book = load_workbook(excel_file)
            sheet = book.active

            # Encontrar la última fila con datos
            last_row = sheet.max_row

            # Agregar los nuevos datos a partir de la siguiente fila
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), last_row + 1):
                for c_idx, value in enumerate(row, 1):
                    sheet.cell(row=r_idx, column=c_idx, value=value)

            # Guardar el archivo
            book.save(excel_file)
        except Exception as e:
            print(f"Error loading or saving Excel file: {e}")
    else:
        # Crear un nuevo archivo de Excel y guardar los datos
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)

# Función para mover el archivo procesado
def move_processed_file(pdf_file, processed_folder):
    if not os.path.exists(processed_folder):
        os.makedirs(processed_folder)
    shutil.move(pdf_file, os.path.join(processed_folder, os.path.basename(pdf_file)))

# Main code
folder_path = 'facturas'
processed_folder = 'facturas_procesadas'
excel_file = 'facturas/facturas.xlsx'

# Obtener todos los archivos PDF en la carpeta
pdf_files = get_files_in_carpetas(folder_path)

# Lista para almacenar los datos extraídos de cada factura
invoice_data_list = []

# Extraer la información de cada PDF
for pdf_file in pdf_files:
    invoice_data = extract_factura_info(pdf_file)
    invoice_data_list.append(invoice_data)

    # Mover el archivo procesado a la carpeta de facturas procesadas
    move_processed_file(pdf_file, processed_folder)

# Crear un DataFrame con la información de todas las facturas
all_invoices = pd.DataFrame(invoice_data_list)

# Verificar si la fecha está presente en el DataFrame
if 'Fecha' not in all_invoices.columns:
    all_invoices['Fecha'] = None  # Si no hay columna de fecha, crearla

# Guardar la información en un archivo Excel sin sobrescribir los datos anteriores
save_to_excel(all_invoices, excel_file)
